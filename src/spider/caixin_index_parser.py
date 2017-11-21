from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
"""
bs4练习，解析财新网首页新闻，导入excel文件
"""


# 获取url下的网页内容
def get_content(url):
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.0'}
    return requests.request(method='get', url=url, headers=header).content


# 解析想要的数据 图片 链接 标题
def parser_html(html):
    result = []
    soup = BeautifulSoup(html, 'html.parser')
    for dl in soup.find_all('dl'):
        news = []
        if dl.find('img'):
            news.append(dl.find('img')['src'])
        else:
            news.append("无图新闻")
        if dl.find('p') and dl.find('p').find('a'):
            news.append(dl.find('p').find('a')['href'])
            news.append(dl.find('p').find('a').string)
        elif dl.find('dt') and dl.find('dt').find('a'):
            news.append(dl.find('dt').find('a')['href'])
            news.append(dl.find('dt').find('a').string)
        elif dl.find('span') and dl.find('span').find('a'):
            news.append(dl.find('span').find('a')['href'])
            news.append(dl.find('div').find('a').string)
        else:
            news.append(None)
            news.append(None)
        result.append(news)
    return result


if __name__ == '__main__':
    url = 'http://www.caixin.com/index.html'
    html = get_content(url)
    result = parser_html(html)
    wb = Workbook()
    ws = wb.active
    ws.title = '新闻列表'
    for index in range(0, result.__len__()):
        col_A = 'A%s' % (index + 1)
        col_B = 'B%s' % (index + 1)
        col_C = 'C%s' % (index + 1)
        ws[col_A] = result[index][0]
        ws[col_B] = result[index][1]
        ws[col_C] = result[index][2]
    wb.save('财新首页新闻.xlsx')
